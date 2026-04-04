"""
Excel → DataModel 変換エンジン

設計原則:
  - LLMは「判断」だけ（構造検出 + 列の意味理解 = 2回のみ）
  - 変換ロジックはすべて決定論的Pythonコード
  - ヘッダーのパターンが不明でも処理できる
  - 5万行でもチャンク処理で対応
"""
import asyncio
import logging
import io
import re
import openpyxl
from datetime import datetime
from typing import Any, Optional
from dataclasses import dataclass, field

from app.services.excel_agents import run_phase1a_search, run_phase1b_global_state, run_phase2_parallel_mapping


logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# データクラス定義
# ═══════════════════════════════════════════════════════════════

@dataclass
class CellMergeInfo:
    """結合セル情報"""
    min_row: int
    min_col: int
    max_row: int
    max_col: int


@dataclass
class PhysicalGrid:
    """正規化済み物理グリッド"""
    rows: list[list[str]]          # 全行データ（結合展開済み）
    total_rows: int
    sheet_name: str
    merge_map: dict = field(default_factory=dict)  # (row,col) -> CellMergeInfo
    bold_rows: set = field(default_factory=set)     # 太字の行番号集合
    bg_rows: set = field(default_factory=set)       # 背景色のある行番号集合


@dataclass
class StructureInfo:
    """LLMによる構造検出結果"""
    header_start: int          # ヘッダー開始行（0-indexed）
    header_end: int            # ヘッダー終了行（inclusive）
    data_start: int            # データ開始行
    blank_row_strategy: str    # "skip" | "group_separator"
    key_column: int            # 機器ID列（0-indexed）
    forward_fill_columns: list[int]  # 空白継承すべき列番号リスト
    pattern: str               # "A"=機器仕様, "B"=星取表, "C"=混合
    year_context: int          # 年度コンテキスト
    implied_hierarchy: dict[str, str] = field(default_factory=dict) # 暗黙の階層推論
    is_target_sheet: bool = True # 探索除外対象かどうか

@dataclass
class ColumnDescriptor:
    """列記述子: 各物理列の意味"""
    col: int                   # 列番号（0-indexed）
    field: str                 # "AssetId" | "AssetName" | "WorkOrderName" | "schedule" | "ignore" | etc.
    month: Optional[int] = None  # scheduleの場合の月（1-12）
    sub: Optional[str] = None    # "plan" | "actual" | "both" | None
    label: str = ""              # 人間が読むラベル（「4月-計画」等）


@dataclass
class ValidationResult:
    """マッピング検証結果"""
    preview_records: list[dict]
    warnings: list[str]
    unmapped_columns: list[int]
    symbol_stats: dict[str, int]
    success: bool


# ═══════════════════════════════════════════════════════════════
# Step 1: 物理グリッド正規化（コードのみ）
# ═══════════════════════════════════════════════════════════════

def normalize_all_physical_grids(file_bytes: bytes) -> list[PhysicalGrid]:
    """
    Excelを読み込み、データが含まれるすべてのシートについて
    LLMに渡せる正規化済みグリッドのリストを生成する。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    grids = []
    
    for sheet in wb.worksheets:
        # 結合情報を記録
        merge_map = {}
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            info = CellMergeInfo(min_row - 1, min_col - 1, max_row - 1, max_col - 1)  # 0-indexed
            for r in range(min_row - 1, max_row):
                for c in range(min_col - 1, max_col):
                    merge_map[(r, c)] = info
        
        # 書式情報を抽出（先頭30行）
        bold_rows = set()
        bg_rows = set()
        for row_idx, row in enumerate(sheet.iter_rows(max_row=min(30, sheet.max_row))):
            for cell in row:
                if cell.font and cell.font.bold:
                    bold_rows.add(row_idx)
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                    bg_rows.add(row_idx)
        
        # 結合セルを展開
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(merged_range))
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    sheet.cell(row=r, column=c, value=top_left_value)
        
        # 全行をテキスト化
        rows = []
        total_rows = 0
        has_data = False
        for row in sheet.iter_rows(values_only=True):
            row_str = [str(cell).strip() if cell is not None else "" for cell in row]
            if any(row_str):
                has_data = True
            rows.append(row_str)
            total_rows += 1
            
        if has_data:
            grids.append(PhysicalGrid(
                rows=rows,
                total_rows=total_rows,
                sheet_name=sheet.title,
                merge_map=merge_map,
                bold_rows=bold_rows,
                bg_rows=bg_rows,
            ))
            
    return grids


# ═══════════════════════════════════════════════════════════════
# Step 2: 構造検出（LLM呼び出し #1）
# ═══════════════════════════════════════════════════════════════

async def detect_structure(grid: PhysicalGrid, filename: str = "unknown.xlsx") -> StructureInfo:
    """フェーズ1-A & フェーズ1-B: エージェントによる構造・全体像推論"""
    try:
        p1a_data = await run_phase1a_search(grid, filename)
        if not p1a_data:
            logger.warning("[detect_structure] P1-A失敗。ヒューリスティック使用。")
            res = _heuristic_structure_detection(grid)
            res._used_fallback = True
            return res
            
        if not p1a_data.get("is_target_sheet", True):
            p1b_data = {"pattern": "Specs Only", "year_context": 0, "implied_hierarchy": {}}
        else:
            p1b_data = await run_phase1b_global_state(grid, filename, p1a_data)
        
        return StructureInfo(
            header_start=p1a_data.get("header_start", 0),
            header_end=p1a_data.get("header_end", 0),
            data_start=p1a_data.get("data_start", 1),
            blank_row_strategy=p1a_data.get("blank_row_strategy", "skip"),
            key_column=p1a_data.get("key_column", 0),
            forward_fill_columns=p1a_data.get("forward_fill_columns", []),
            pattern=p1b_data.get("pattern", "Specs Only"),
            year_context=p1b_data.get("year_context", 0),
            implied_hierarchy=p1b_data.get("implied_hierarchy", {}),
            is_target_sheet=p1a_data.get("is_target_sheet", True)
        )
    except Exception as e:
        logger.error("[detect_structure] エラー: %s", e)
        res = _heuristic_structure_detection(grid)
        res._used_fallback = True
        return res

def _heuristic_structure_detection(grid: PhysicalGrid) -> StructureInfo:
    """
    LLMが失敗した場合のヒューリスティックフォールバック。
    太字/背景色行をヘッダー候補、その次の非空行をデータ開始とする。
    """
    # ヘッダー候補: 太字+背景色のある行
    header_candidates = sorted(grid.bold_rows & grid.bg_rows)
    if not header_candidates:
        header_candidates = sorted(grid.bold_rows | grid.bg_rows)
    
    if header_candidates:
        # 連続するヘッダー行を検出
        header_start = header_candidates[0]
        header_end = header_start
        for h in header_candidates[1:]:
            if h == header_end + 1:
                header_end = h
            else:
                break
        data_start = header_end + 1
    else:
        # 書式情報なし: 最初の行をヘッダーと仮定
        header_start = 0
        header_end = 0
        data_start = 1
    
    # key_column: 最初の非空列
    key_column = 0
    
    # パターン判定: シンボルが含まれるか
    symbols = set("○●◎△×☆★◇◆")
    has_symbols = False
    for i in range(data_start, min(data_start + 10, len(grid.rows))):
        for cell in grid.rows[i]:
            if cell.strip() in symbols:
                has_symbols = True
                break
    
    # 年度推定
    year_context = 0
    for row in grid.rows[:5]:
        for cell in row:
            m = re.search(r'(20\d{2})', cell)
            if m:
                year_context = int(m.group(1))
                break
        if year_context:
            break
    
    logger.info(
        "[heuristic] header=%d-%d, data=%d, pattern=%s, year=%d",
        header_start, header_end, data_start,
        "B" if has_symbols else "A", year_context,
    )
    return StructureInfo(
        header_start=header_start,
        header_end=header_end,
        data_start=data_start,
        blank_row_strategy="skip",
        key_column=key_column,
        forward_fill_columns=[key_column],
        pattern="Schedule Matrix" if has_symbols else "Specs Only",
        year_context=year_context,
    )


# ═══════════════════════════════════════════════════════════════
# Step 3: 列記述子生成（LLM呼び出し #2）
# ═══════════════════════════════════════════════════════════════

async def generate_column_descriptors(
    grid: PhysicalGrid,
    structure: StructureInfo,
) -> tuple[list[ColumnDescriptor], bool]:
    """フェーズ2: 並列エージェントによる列マッピング"""
    try:
        descriptors = await run_phase2_parallel_mapping(grid, structure)
        if not descriptors:
            logger.warning("[generate_column_descriptors] descriptors空。ヒューリスティック使用。")
            return _heuristic_column_descriptors(grid, structure), True
        return descriptors, False
    except Exception as e:
        logger.error("[generate_column_descriptors] エラー: %s", e)
        return _heuristic_column_descriptors(grid, structure), True

def _heuristic_column_descriptors(
    grid: PhysicalGrid,
    structure: StructureInfo,
) -> list[ColumnDescriptor]:
    """
    LLMが失敗した場合のヒューリスティック列記述子生成。
    ヘッダー行の内容から列の意味を推定する。
    """
    descriptors = []
    header_rows = grid.rows[structure.header_start : structure.header_end + 1]
    
    if not header_rows:
        return descriptors
    
    # 最後のヘッダー行を主キーにする
    main_header = header_rows[-1] if len(header_rows) > 0 else []
    
    # 名称キーワード
    id_keywords = ["番号", "No", "TAG", "ID", "コード", "機器番号"]
    name_keywords = ["名", "名称", "機器台帳", "設備名"]
    year_pattern = re.compile(r'20\d{2}')
    month_pattern = re.compile(r'(\d{1,2})月')
    
    for col_idx in range(len(main_header)):
        # 全ヘッダー行のこの列の値を収集
        col_headers = []
        for h_row in header_rows:
            if col_idx < len(h_row) and h_row[col_idx].strip():
                col_headers.append(h_row[col_idx].strip())
        
        header_text = " ".join(col_headers)
        
        if not header_text:
            descriptors.append(ColumnDescriptor(col=col_idx, field="ignore", label="空ヘッダー"))
            continue
        
        # ID列判定
        if any(kw in header_text for kw in id_keywords):
            descriptors.append(ColumnDescriptor(col=col_idx, field="AssetId", label=header_text))
        # 名称列判定
        elif any(kw in header_text for kw in name_keywords):
            descriptors.append(ColumnDescriptor(col=col_idx, field="AssetName", label=header_text))
        # 年度列判定（2021, 2022等）
        elif year_pattern.search(header_text):
            descriptors.append(ColumnDescriptor(
                col=col_idx, field="schedule", sub="both",
                label=header_text,
            ))
        # 月列判定
        elif month_pattern.search(header_text):
            m = month_pattern.search(header_text)
            descriptors.append(ColumnDescriptor(
                col=col_idx, field="schedule", month=int(m.group(1)), sub="both",
                label=header_text,
            ))
        else:
            descriptors.append(ColumnDescriptor(col=col_idx, field="Attribute", label=header_text))
    
    logger.info("[heuristic_columns] %d列の記述子を生成", len(descriptors))
    return descriptors


# ═══════════════════════════════════════════════════════════════
# Step 4: 凡例検出（コードのみ）
# ═══════════════════════════════════════════════════════════════

# 保全業界の標準的なシンボルパターン
DEFAULT_SYMBOL_MAPPING = {
    "○": "planned",
    "●": "actual",
    "◎": "planned_and_actual",
    "△": "partial",
    "×": "cancelled",
    "☆": "planned",
    "★": "actual",
}

def detect_legend(grid: PhysicalGrid) -> dict[str, str]:
    """
    Excel内の凡例を探索する。見つからなければデフォルト値を返す。
    """
    legend_keywords = ["凡例", "記号", "マーク", "legend", "Legend", "記号説明"]
    
    for i, row in enumerate(grid.rows):
        for j, cell in enumerate(row):
            for keyword in legend_keywords:
                if keyword in cell:
                    # 凡例キーワードを見つけた → 周辺セルからマッピングを抽出
                    mapping = _extract_legend_from_neighborhood(grid, i, j)
                    if mapping:
                        logger.info("[detect_legend] 凡例を検出: row=%d, col=%d, mapping=%s", i, j, mapping)
                        return mapping
    
    logger.info("[detect_legend] 凡例未検出、デフォルト値を使用")
    return dict(DEFAULT_SYMBOL_MAPPING)


def _extract_legend_from_neighborhood(grid: PhysicalGrid, row: int, col: int) -> dict[str, str]:
    """凡例キーワード周辺からシンボル→意味のマッピングを抽出する"""
    mapping = {}
    symbol_pattern = re.compile(r'^[○●◎△×☆★◇◆□■▲▼]$')
    
    # 下方向に探索（凡例は下に続くことが多い）
    for r in range(row, min(row + 15, len(grid.rows))):
        row_data = grid.rows[r]
        for c in range(max(0, col - 2), min(len(row_data), col + 6)):
            cell = row_data[c].strip()
            if symbol_pattern.match(cell):
                # 次のセルに意味がある
                if c + 1 < len(row_data) and row_data[c + 1].strip():
                    meaning = row_data[c + 1].strip()
                    # 意味をDataModelのフラグに変換
                    if any(k in meaning for k in ["計画", "予定", "plan"]):
                        mapping[cell] = "planned"
                    elif any(k in meaning for k in ["実施", "実績", "完了", "actual", "done"]):
                        mapping[cell] = "actual"
                    elif any(k in meaning for k in ["中止", "cancel"]):
                        mapping[cell] = "cancelled"
                    else:
                        mapping[cell] = "planned"  # デフォルト
    
    return mapping if mapping else None


# ═══════════════════════════════════════════════════════════════
# Step 5: マッピング検証（コードのみ）
# ═══════════════════════════════════════════════════════════════

def validate_and_preview(
    grid: PhysicalGrid,
    structure: StructureInfo,
    descriptors: list[ColumnDescriptor],
    symbol_mapping: dict[str, str],
) -> ValidationResult:
    """サンプル行を使って変換結果をプレビューする"""
    preview_records = []
    warnings = []
    symbol_stats = {}
    
    # 使われていない列を検出
    used_cols = {d.col for d in descriptors if d.field != "ignore"}
    total_cols = len(grid.rows[0]) if grid.rows else 0
    unmapped_columns = [c for c in range(total_cols) if c not in used_cols and c < total_cols]
    
    # サンプル5行を変換
    sample_count = 0
    prev_key_value = ""
    for i in range(structure.data_start, min(structure.data_start + 30, len(grid.rows))):
        row = grid.rows[i]
        if not any(cell.strip() for cell in row):
            continue  # 空白行スキップ
        
        # forward fill
        key_col = structure.key_column
        if key_col < len(row):
            if row[key_col].strip():
                prev_key_value = row[key_col].strip()
            else:
                row = list(row)
                row[key_col] = prev_key_value
        
        record = _convert_single_row(row, descriptors, symbol_mapping, structure.year_context)
        if record:
            preview_records.append(record)
            # シンボル統計
            for d in descriptors:
                if d.field == "schedule" and d.col < len(row):
                    sym = row[d.col].strip()
                    if sym:
                        symbol_stats[sym] = symbol_stats.get(sym, 0) + 1
        
        sample_count += 1
        if sample_count >= 5:
            break
    
    # 検証警告
    if not preview_records:
        warnings.append("サンプル行からレコードを生成できませんでした")
    
    asset_desc = [d for d in descriptors if d.field == "AssetId"]
    if not asset_desc:
        warnings.append("機器ID列が特定されていません")
    
    schedule_desc = [d for d in descriptors if d.field == "schedule"]
    if "Schedule" in structure.pattern and not schedule_desc:
        warnings.append("星取表パターンですが、スケジュール列が特定されていません")
    
    return ValidationResult(
        preview_records=preview_records,
        warnings=warnings,
        unmapped_columns=unmapped_columns,
        symbol_stats=symbol_stats,
        success=len(warnings) == 0 or len(preview_records) > 0,
    )


# ═══════════════════════════════════════════════════════════════
# Step 6: チャンク変換（コードのみ）
# ═══════════════════════════════════════════════════════════════

def convert_chunk(
    grid: PhysicalGrid,
    structure: StructureInfo,
    descriptors: list[ColumnDescriptor],
    symbol_mapping: dict[str, str],
    start_row: int,
    chunk_size: int = 500,
) -> dict[str, Any]:
    """
    確定済みの構造情報+列記述子+シンボルマッピングで
    決定論的にDataModel JSONに変換する。LLM呼び出しゼロ。
    """
    assets = {}       # key: asset_id
    work_orders = {}   # key: wo_name
    wo_lines = []
    
    end_row = min(start_row + chunk_size, len(grid.rows))
    processed_count = 0
    error_rows = []
    prev_key_values = {}  # col -> last non-empty value
    
    for i in range(start_row, end_row):
        row = grid.rows[i]
        processed_count += 1
        
        # ヘッダー行以前はスキップ
        if i <= structure.header_end:
            continue
        
        # 空白行処理
        if not any(cell.strip() for cell in row):
            if structure.blank_row_strategy == "skip":
                continue
            else:
                # group_separator: forward_fill をリセット
                prev_key_values.clear()
                continue
        
        # forward fill
        row = list(row)
        for fcol in structure.forward_fill_columns:
            if fcol < len(row):
                if row[fcol].strip():
                    prev_key_values[fcol] = row[fcol].strip()
                elif fcol in prev_key_values:
                    row[fcol] = prev_key_values[fcol]
        
        try:
            record = _convert_single_row(
                row, 
                descriptors, 
                symbol_mapping, 
                structure.year_context, 
                structure.implied_hierarchy,
                grid.sheet_name,
                i
            )
            if record:
                _merge_record_into_model(record, assets, work_orders, wo_lines)
        except Exception as e:
            error_rows.append({"row": i, "error": str(e)})
    
    return {
        "assets": list(assets.values()),
        "work_orders": list(work_orders.values()),
        "work_order_lines": wo_lines,
        "processed_count": processed_count,
        "error_rows": error_rows,
        "is_done": end_row >= len(grid.rows),
    }


def _convert_single_row(
    row: list[str],
    descriptors: list[ColumnDescriptor],
    symbol_mapping: dict[str, str],
    year_context: int,
    implied_hierarchy: dict[str, str] = None,
    sheet_name: str = "Unknown",
    row_idx: int = 0
) -> dict | None:
    """1行をDataModelレコードに変換する"""
    record = {
        "asset_id": "",
        "asset_name": "",
        "wo_name": "",
        "classification": "",
        "hierarchyPath": {},
        "specs": {},
        "schedule_entries": [],
    }
    
    hierarchy_from_cols = {}
    
    for d in descriptors:
        if d.col >= len(row) or d.field == "ignore":
            continue
        val = row[d.col].strip()
        if not val:
            continue
        
        if d.field == "AssetId":
            record["asset_id"] = val
        elif d.field == "AssetName":
            record["asset_name"] = val
        elif d.field == "WorkOrderName":
            record["wo_name"] = val
        elif d.field == "Classification":
            record["classification"] = val
        elif d.field == "Hierarchy1":
            hierarchy_from_cols["第1階層"] = val
        elif d.field == "Hierarchy2":
            hierarchy_from_cols["第2階層"] = val
        elif d.field == "Hierarchy3":
            hierarchy_from_cols["第3階層"] = val
        elif d.field == "Attribute" or d.field in ("Manufacturer", "Model", "Location", "Remarks"):
            key = d.label if d.label else d.field.lower()
            record["specs"][key] = val
        elif d.field == "schedule" and d.month is not None:
            meaning = symbol_mapping.get(val, "")
            year = year_context if year_context > 0 else datetime.now().year
            # 4月始まりの場合、1-3月は翌年
            actual_year = year + 1 if d.month <= 3 else year
            entry = {
                "month": d.month,
                "year": actual_year,
                "symbol": val,
                "meaning": meaning,
                "date": f"{actual_year}-{d.month:02d}-01",
                "sub": d.sub or "both",
            }
            record["schedule_entries"].append(entry)
    
    # 機器IDの自動補完 (フロントエンドのバリデーションエラー防止)
    if not record["asset_id"]:
        if record["asset_name"]:
            import urllib.parse
            safe_name = urllib.parse.quote(record["asset_name"])[:20]
            record["asset_id"] = f"AUTO-{safe_name}-{row_idx}"
        else:
            record["asset_id"] = f"AUTO-{sheet_name}-{row_idx:04d}"
            record["asset_name"] = f"不明機器 ({sheet_name} {row_idx}行目)"
    
    # HierarchyPathの適用
    # 優先順位: 1. 列から抽出したもの 2. LLM推論のimplied_hierarchy 3. 無ければ "未設定"
    if hierarchy_from_cols:
        record["hierarchyPath"] = hierarchy_from_cols
    elif implied_hierarchy and len(implied_hierarchy) > 0:
        record["hierarchyPath"] = implied_hierarchy
    else:
        record["hierarchyPath"] = {"分類": "インポート未設定"}
    
    # 最低限のデータがなければスキップ
    if not record["asset_id"] and not record["wo_name"]:
        return None
    
    return record


def _merge_record_into_model(
    record: dict,
    assets: dict,
    work_orders: dict,
    wo_lines: list,
):
    """変換済みレコードをDataModelに統合する"""
    asset_id = record["asset_id"]
    wo_name = record["wo_name"]
    
    # Asset
    if asset_id:
        if asset_id not in assets:
            assets[asset_id] = {
                "id": asset_id,
                "name": record["asset_name"] or asset_id,
                "hierarchyPath": dict(record["hierarchyPath"]),
            }
            # specsを追加
            for k, v in record["specs"].items():
                if v:
                    assets[asset_id][k] = v
        else:
            # 既存のAssetにspecsをマージ
            for k, v in record["specs"].items():
                if v and k not in assets[asset_id]:
                    assets[asset_id][k] = v
    
    # WorkOrder
    if wo_name and wo_name not in work_orders:
        wo_id = f"WO-{len(work_orders) + 1:04d}"
        work_orders[wo_name] = {
            "id": wo_id,
            "name": wo_name,
            "Classification": record["classification"] or "05",
        }
    
    # WorkOrderLines (from schedule entries)
    wo_id = work_orders.get(wo_name, {}).get("id", "")
    for entry in record["schedule_entries"]:
        is_planned = entry["meaning"] in ("planned", "planned_and_actual", "partial", "")
        is_actual = entry["meaning"] in ("actual", "planned_and_actual")
        
        if entry["sub"] == "plan":
            is_actual = False
            is_planned = True
        elif entry["sub"] == "actual":
            is_planned = False
            is_actual = True
        
        wol = {
            "id": f"WOL-IMP-{len(wo_lines) + 1:06d}",
            "AssetId": asset_id,
            "WorkOrderId": wo_id,
            "WorkOrderName": wo_name,
            "PlanScheduleStart": entry["date"] if is_planned else "",
            "ActualScheduleStart": entry["date"] if is_actual else "",
            "Planned": is_planned,
            "Symbol": entry["symbol"],
        }
        wo_lines.append(wol)


# ═══════════════════════════════════════════════════════════════
# 高レベルAPI（外部から呼ばれるエントリポイント）
# ═══════════════════════════════════════════════════════════════

async def _process_single_sheet(grid: PhysicalGrid, filename: str) -> dict[str, Any]:
    """単一シートの構造解析 + 列マッピング + 凡例検出 + 検証プレビューを行う。"""
    logger.info("[analyze_excel_full] シート処理開始: %s (%d行)", grid.sheet_name, grid.total_rows)
    fallback_warnings = []

    # Step 2: LLMによる構造検出
    structure = await detect_structure(grid, filename)
    if getattr(structure, '_used_fallback', False):
        fallback_warnings.append("AI解析が利用できなかったため構造検出に簡易解析を使用しました")
    logger.info(
        "[analyze_excel_full] 構造検出完了: header=%d-%d, data_start=%d, pattern=%s",
        structure.header_start, structure.header_end, structure.data_start, structure.pattern,
    )

    # Step 3: LLMによる列記述子生成
    if getattr(structure, 'is_target_sheet', True):
        descriptors, desc_used_fallback = await generate_column_descriptors(grid, structure)
        if desc_used_fallback:
            fallback_warnings.append("AI解析が利用できなかったため列マッピングに簡易解析を使用しました")
        logger.info("[analyze_excel_full] 列記述子生成完了: %d列", len(descriptors))
    else:
        logger.info("[analyze_excel_full] is_target_sheet=FalseのためPhase2 LLMマッピングをスキップ")
        descriptors = []

    # Step 4: 凡例検出
    symbol_mapping = detect_legend(grid)

    # Step 5: 検証プレビュー
    validation = validate_and_preview(grid, structure, descriptors, symbol_mapping)
    # フォールバック警告を検証結果のwarningsに追加
    validation.warnings.extend(fallback_warnings)
    logger.info("[analyze_excel_full] 検証完了: records=%d", len(validation.preview_records))

    return {
        "grid": grid,
        "structure": structure,
        "descriptors": descriptors,
        "symbol_mapping": symbol_mapping,
        "validation": validation,
        "total_rows": grid.total_rows,
        "sheet_name": grid.sheet_name,
        "filename": filename,
    }


async def analyze_excel_full(file_bytes: bytes, filename: str = "unknown.xlsx", session_id: str = None) -> list[dict[str, Any]]:
    """
    Phase 1-2: 構造解析 + マッピング生成 + 検証プレビュー
    全シート（データがあるもの）に対して並列処理し、リストで返す。
    """
    logger.info("[analyze_excel_full] 開始: %s", filename)

    # Step 1: 物理グリッド正規化（複数シート）
    grids = normalize_all_physical_grids(file_bytes)
    logger.info("[analyze_excel_full] グリッド正規化完了: %dシート", len(grids))

    # 各シートを順次処理（Local LLMのキュー溢れを防ぐため直列実行）
    from app.services.session_manager import session_manager
    
    results = []
    for grid in grids:
        if session_id:
            session = session_manager.get_session(session_id)
            if getattr(session, 'is_cancelled', False):
                logger.warning(f"[analyze_excel_full] セッション {session_id} により処理がキャンセルされました")
                raise asyncio.CancelledError("User cancelled the import process")
                
        results.append(await _process_single_sheet(grid, filename))

    return results


def execute_chunk_conversion(
    analysis_results: list[dict[str, Any]],
    chunk_size: int = 500,
) -> dict[str, Any]:
    """
    Phase 3: チャンク変換。LLM呼び出しゼロ。複数シートを全て変換して統合する。
    """
    assets = {}
    work_orders = {}
    wo_lines = []
    error_rows = []
    processed_count = 0
    
    for analysis_result in analysis_results:
        grid = analysis_result["grid"]
        structure = analysis_result["structure"]
        descriptors = analysis_result["descriptors"]
        symbol_mapping = analysis_result["symbol_mapping"]
        
        # data_start以降から処理
        actual_start = structure.data_start
        
        # シート単位のチャンク処理
        total_rows = grid.total_rows
        current_row = actual_start
        while current_row < total_rows:
            chunk_result = convert_chunk(
                grid=grid,
                structure=structure,
                descriptors=descriptors,
                symbol_mapping=symbol_mapping,
                start_row=current_row,
                chunk_size=chunk_size,
            )
            
            for asset in chunk_result["assets"]:
                aid = asset["id"]
                if aid not in assets:
                    assets[aid] = dict(asset)
                else:
                    # マージ
                    for k, v in asset.items():
                        if k not in ["id", "name", "hierarchyPath"] and v and k not in assets[aid]:
                            assets[aid][k] = v
            for wo in chunk_result["work_orders"]:
                work_orders[wo["name"]] = wo
            wo_lines.extend(chunk_result["work_order_lines"])
            error_rows.extend([{"sheet": grid.sheet_name, **err} for err in chunk_result["error_rows"]])
            processed_count += chunk_result["processed_count"]
            current_row += chunk_size
            
    return {
        "assets": list(assets.values()),
        "work_orders": list(work_orders.values()),
        "work_order_lines": wo_lines,
        "processed_count": processed_count,
        "error_rows": error_rows,
        "is_done": True,
    }
